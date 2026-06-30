# Dashboard SCIAN — Peibo
# Ejecutar desde PowerShell:
# & "C:/Users/Laura Valentina/AppData/Local/Programs/Python/Python312/python.exe" -m streamlit run "C:/Users/Laura Valentina/OneDrive/Documents/Proyecto Data CREDITRUST/Dashboard para PEIBO.py"

import io, json, os, re
from collections import defaultdict
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# ══════════════════════════════════════════════════════════════════════════════
# SISTEMA DE DISEÑO — Índigo Pizarra · Acento Oro · 4 niveles de elevación
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="SCIAN Analytics | Peibo", page_icon="🏢", layout="wide")

C = {
    # Elevación — cada nivel un paso más claro que el anterior
    "bg":          "#030913",   # nivel 0 — página
    "surface":     "#08121F",   # nivel 1 — tarjeta
    "surface2":    "#0E1E35",   # nivel 2 — panel / hover / insight box
    "surface3":    "#152844",   # nivel 3 — popover / tooltip
    "border":      "#162844",
    "border2":     "#1F3C66",
    "border3":     "#2A4D80",
    # Acento principal — oro apagado
    "gold":        "#C9A84C",
    "gold_dim":    "#7A6128",
    "gold_glow":   "#C9A84C1A",
    # Actividad principal — azul
    "blue":        "#4F94F8",
    "blue_dim":    "#12306B",
    "blue_mid":    "#2A5DC4",
    # Actividad secundaria — índigo-violeta
    "indigo":      "#7C93F5",
    "indigo_dim":  "#1A1F5E",
    "indigo_mid":  "#4F62D9",
    "indigo2":     "#6366F1",
    # Estados / semántica
    "green":       "#10B981",
    "warn":        "#F59E0B",
    "danger":      "#EF4444",
    "purple":      "#A78BFA",
    # Texto — 3 niveles de contraste, todos AA o mejor
    "text_hi":     "#F4F7FF",   # 15.8:1 sobre surface
    "text_md":     "#94AABF",   # 7.2:1 sobre surface
    "text_lo":     "#5A7499",   # 4.6:1 sobre surface — cumple AA
    "text_dim":    "#2C4260",   # solo decorativo, nunca para texto legible
    "grid":        "#0A1829",
}

st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

  /* ── Base ── */
  .stApp {{ background-color:{C['bg']}; font-family:'Inter',sans-serif; }}
  section[data-testid="stSidebar"] {{
      background-color:{C['surface']};
      border-right:1px solid {C['border']};
  }}
  #MainMenu, footer, header {{ visibility:hidden; }}

  /* ── Sidebar ── */
  section[data-testid="stSidebar"] .stMarkdown p {{ color:{C['text_md']}; font-size:.85rem; }}
  section[data-testid="stSidebar"] .stMarkdown b {{ color:{C['gold']}; }}

  /* ── Page header ── */
  .ph-wrap {{
      border-bottom:1px solid {C['border']};
      padding-bottom:20px; margin-bottom:24px;
  }}
  .ph-eyebrow {{
      font-size:.68rem; font-weight:600; text-transform:uppercase;
      letter-spacing:.14em; color:{C['gold']}; margin:0 0 6px 0;
  }}
  .ph-title {{
      font-size:1.75rem; font-weight:700; color:{C['text_hi']};
      margin:0 0 4px 0; letter-spacing:-.025em;
  }}
  .ph-sub {{ font-size:.9rem; color:{C['text_md']}; margin:0; }}

  /* ── KPI cards — nivel 1, borde superior por color semántico ── */
  .kpi-grid {{
      display:grid; grid-template-columns:repeat(6,1fr);
      gap:10px; margin-bottom:24px;
  }}
  .kpi-card {{
      background:{C['surface']};
      border:1px solid {C['border']};
      border-radius:10px;
      padding:16px 16px 13px;
      position:relative; overflow:hidden;
      transition:border-color .2s;
  }}
  .kpi-card:hover {{ border-color:{C['border2']}; }}
  .kpi-card::after {{ content:''; position:absolute; top:0; left:0; right:0; height:2px; }}
  .kc-gold::after   {{ background:{C['gold']}; }}
  .kc-blue::after   {{ background:{C['blue']}; }}
  .kc-indigo::after {{ background:{C['indigo']}; }}
  .kc-green::after  {{ background:{C['green']}; }}
  .kc-warn::after   {{ background:{C['warn']}; }}
  .kc-slate::after  {{ background:{C['text_lo']}; }}
  .kc-purple::after {{ background:{C['purple']}; }}
  .kc-danger::after {{ background:{C['danger']}; }}
  .kpi-lbl {{ font-size:.65rem; text-transform:uppercase; letter-spacing:.09em; color:{C['text_lo']}; margin:0 0 7px 0; }}
  .kpi-val {{ font-size:1.5rem; font-weight:700; color:{C['text_hi']}; line-height:1; margin:0 0 5px 0; letter-spacing:-.02em; }}
  .kpi-sub {{ font-size:.78rem; color:{C['text_md']}; margin:0; }}

  /* ── Tabs ── */
  .stTabs [data-baseweb="tab-list"] {{
      gap:2px; background:{C['surface']}; border-radius:8px;
      padding:4px; border:1px solid {C['border']}; margin-bottom:24px;
  }}
  .stTabs [data-baseweb="tab"] {{
      border-radius:6px; padding:9px 22px;
      font-size:.85rem; font-weight:500; letter-spacing:.01em;
      color:{C['text_md']}; background:transparent; border:none !important;
  }}
  .stTabs [aria-selected="true"] {{
      background:{C['surface2']} !important;
      color:{C['text_hi']} !important;
      border-bottom:2px solid {C['gold']} !important;
      font-weight:600;
  }}
  .stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] {{ display:none; }}

  /* ── Section titles ── */
  .sec-title {{
      font-size:1rem; font-weight:600; color:{C['text_hi']};
      margin:32px 0 3px 0; padding-bottom:10px;
      border-bottom:1px solid {C['border']};
      display:flex; align-items:center; gap:8px;
  }}
  .sec-title::before {{
      content:''; display:inline-block; width:3px; height:14px;
      border-radius:2px; background:{C['gold']}; flex-shrink:0;
  }}
  .sec-desc {{ font-size:.8rem; color:{C['text_md']}; margin:6px 0 18px 0; line-height:1.55; }}

  /* ── Chart labels ── */
  /* Streamlit envuelve st.container(border=True) en un div con data-testid="stVerticalBlockBorderWrapper".
     Restyle nativo para que coincida con el sistema de diseño en vez del borde gris por defecto. */
  div[data-testid="stVerticalBlockBorderWrapper"] {{
      margin-bottom:14px;
  }}
  div[data-testid="stVerticalBlockBorderWrapper"] > div {{
      background:{C['surface']}; border:1px solid {C['border']} !important;
      border-radius:12px !important; padding:6px 4px 10px 4px;
  }}
  .ct {{ font-size:.85rem; font-weight:600; color:{C['text_hi']}; margin:0 0 3px 0; }}
  .cs {{ font-size:.78rem; color:{C['text_md']}; margin:0 0 4px 0; line-height:1.5; }}
  .cs2 {{ font-size:.72rem; color:{C['text_lo']}; margin:0 0 14px 0; }}
  .cs em {{ color:{C['blue']}; font-style:normal; font-weight:500; }}
  .cs i {{ color:{C['indigo']}; font-style:normal; font-weight:500; }}

  /* ── Insight box — nivel 2 ── */
  .insight-box {{
      background:{C['gold_glow']}; border:1px solid {C['gold_dim']};
      border-radius:8px; padding:11px 16px;
      font-size:.82rem; color:{C['text_md']}; margin-bottom:18px;
  }}
  .insight-box b {{ color:{C['gold']}; }}

  /* ── Legend chips ── */
  .legend-row {{ display:flex; gap:12px; margin-bottom:14px; flex-wrap:wrap; }}
  .legend-chip {{
      display:flex; align-items:center; gap:6px;
      background:{C['surface2']}; border:1px solid {C['border2']};
      border-radius:8px; padding:6px 13px; font-size:.78rem; color:{C['text_md']};
  }}
  .legend-dot {{ width:10px; height:10px; border-radius:2px; flex-shrink:0; }}

  /* ── Misc ── */
  .stDataFrame {{ border-radius:10px; overflow:hidden; }}
  hr.div {{ border:none; border-top:1px solid {C['border']}; margin:24px 0; }}
</style>
""", unsafe_allow_html=True)

# ── Plotly base ───────────────────────────────────────────────────────────────
GR = dict(gridcolor=C["grid"], linecolor=C["grid"], zerolinecolor=C["grid"])
PL = dict(
    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
    font=dict(color=C["text_md"], family="Inter,sans-serif", size=11),
    legend=dict(bgcolor="rgba(0,0,0,0)", bordercolor="rgba(0,0,0,0)"),
    margin=dict(l=10, r=20, t=20, b=12),
    hoverlabel=dict(bgcolor=C["surface3"], bordercolor=C["border3"],
                    font=dict(color=C["text_hi"], size=12)),
)

def lay(fig, **kw):
    fig.update_layout(**{**PL, **kw})
    return fig

def pch(fig):
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

def sec(title, desc=""):
    st.markdown(f'<p class="sec-title">{title}</p>', unsafe_allow_html=True)
    if desc:
        st.markdown(f'<p class="sec-desc">{desc}</p>', unsafe_allow_html=True)

def chart_card():
    """Devuelve un st.container con borde nativo — envuelve widgets (incl. Plotly) sin romper el DOM."""
    return st.container(border=True)

def chart_header(title, sub="", sub2=""):
    s  = f'<p class="cs">{sub}</p>' if sub else ""
    s2 = f'<p class="cs2">{sub2}</p>' if sub2 else ""
    st.markdown(f'<p class="ct">{title}</p>{s}{s2}', unsafe_allow_html=True)

def legend_row(items):
    chips = "".join(
        f'<div class="legend-chip"><span class="legend-dot" style="background:{color};"></span>{label}</div>'
        for label, color in items
    )
    st.markdown(f'<div class="legend-row">{chips}</div>', unsafe_allow_html=True)

def hr():
    st.markdown("<hr class='div'>", unsafe_allow_html=True)

def kpi_card(val, label, sub, cls):
    return f"""
    <div class="kpi-card {cls}">
      <p class="kpi-lbl">{label}</p>
      <p class="kpi-val">{val}</p>
      <p class="kpi-sub">{sub}</p>
    </div>"""

def kpi_grid(cards_html, cols=6):
    st.markdown(
        f'<div class="kpi-grid" style="grid-template-columns:repeat({cols},1fr);">{cards_html}</div>',
        unsafe_allow_html=True)

def bar_h(df_in, x_col, y_col, nom_col, pct_col, color_lo, color_hi, height=None):
    df_in = df_in.copy()
    df_in[y_col] = df_in[y_col].astype(str)
    h = height or max(260, len(df_in) * 34 + 60)
    fig = go.Figure(go.Bar(
        x=df_in[x_col], y=df_in[y_col], orientation="h",
        marker=dict(color=df_in[x_col],
                    colorscale=[[0, color_lo], [1, color_hi]],
                    cmin=0, cmax=df_in[x_col].max(),
                    line=dict(width=0)),
        customdata=list(zip(df_in[nom_col], df_in[pct_col])),
        text=df_in[pct_col].astype(str) + "%",
        textposition="outside",
        textfont=dict(color=C["text_lo"], size=10),
        hovertemplate=(
            "<b>%{y}</b><br>%{customdata[0]}<br>"
            "Clientes: <b>%{x:,}</b> · %{customdata[1]}%"
            "<extra></extra>"
        ),
    ))
    lay(fig,
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        yaxis=dict(type="category", categoryorder="total ascending",
                   showgrid=False, tickfont=dict(color=C["text_md"], size=11)),
        height=h)
    return fig

def get_combos(df_ppal_in, df_sec_in, top=12, allowed_subs=None):
    pm = df_ppal_in[["cliente","sub_cod","sub_nom"]].rename(
        columns={"sub_cod":"ppal_cod","sub_nom":"ppal_nom"})
    sm = df_sec_in[["cliente","sub_cod","sub_nom"]].rename(
        columns={"sub_cod":"sec_cod","sub_nom":"sec_nom"})
    merged = pm.merge(sm, on="cliente")
    if merged.empty: return pd.DataFrame()
    out = (merged.groupby(["ppal_cod","ppal_nom","sec_cod","sec_nom"])
           .size().reset_index(name="Clientes")
           .sort_values("Clientes", ascending=False))
    if allowed_subs is not None:
        out = out[out["ppal_cod"].isin(allowed_subs) & out["sec_cod"].isin(allowed_subs)]
    return out.head(top)


# ══════════════════════════════════════════════════════════════════════════════
# CARGA DE DATOS Y EXCEL
# ══════════════════════════════════════════════════════════════════════════════

CARPETA = os.path.dirname(os.path.abspath(__file__))

def expandir(cod_raw):
    cod = str(cod_raw).strip()
    m = re.match(r'^(\d+)-(\d+)$', cod)
    return [str(n) for n in range(int(m.group(1)), int(m.group(2))+1)] if m else [cod]

@st.cache_data
def cargar_catalogo():
    ruta = os.path.join(CARPETA, "scian_2023_categorias_y_productos.xlsx")
    if not os.path.exists(ruta): return {}, {}
    try:
        ds = pd.read_excel(ruta, sheet_name="SECTOR",    usecols=[0,1], header=0, dtype=str)
        db = pd.read_excel(ruta, sheet_name="SUBSECTOR", usecols=[0,1], header=0, dtype=str)
    except: return {}, {}
    def td(df):
        d, cols = {}, df.columns.tolist()
        for _, r in df.iterrows():
            c = str(r[cols[0]] or "").strip(); n = str(r[cols[1]] or "").strip()
            if c and n and c.lower() not in ("nan","none"):
                for x in expandir(c): d[x] = n
        return d
    return td(ds), td(db)

@st.cache_data
def cargar_json():
    ruta = os.path.join(CARPETA, "datos_companyPEIBO_subsector.json")
    if not os.path.exists(ruta): return []
    with open(ruta, encoding="utf-8") as f: raw = json.load(f)
    if isinstance(raw, list): return raw
    for v in raw.values():
        if isinstance(v, list): return v
    return []

@st.cache_data
def build_df(clientes_json, cat_sec, cat_sub):
    rows = []
    for e in json.loads(clientes_json):
        cliente = e.get("Cliente")
        subs = sorted(e.get("ScianSubsector",[]), key=lambda x: x.get("Percentage",0), reverse=True)
        for i, s in enumerate(subs):
            cod = str(s.get("CodigoSubsectorSCIAN") or "N/A").strip()
            sc  = cod[:2] if len(cod) >= 2 else "N/A"
            rows.append({
                "cliente":  cliente,
                "tipo_act": "Principal" if i == 0 else "Secundaria",
                "sub_cod":  cod,
                "sub_nom":  cat_sub.get(cod) or str(s.get("NameSubsectorSCIAN") or "Sin nombre"),
                "sec_cod":  sc,
                "sec_nom":  cat_sec.get(sc, "Sin clasificar"),
                "pct":      s.get("Percentage", 0),
            })
    return pd.DataFrame(rows)

@st.cache_data
def generar_excel(clientes_json, cat_sec_items, cat_sub_items):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cat_sec = dict(cat_sec_items); cat_sub = dict(cat_sub_items)
    clientes = json.loads(clientes_json)
    AZUL="1F4E79"; CLAR="D6E4F0"

    def borde():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)
    def enc(cell, bg=AZUL, fg="FFFFFF", sz=10):
        cell.font=Font(bold=True,color=fg,size=sz)
        cell.fill=PatternFill("solid",fgColor=bg)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=borde()
    def cel(cell, bold=False, center=False, bg=None, sz=10):
        cell.font=Font(bold=bold,size=sz)
        cell.alignment=Alignment(horizontal="center" if center else "left",vertical="center")
        cell.border=borde()
        if bg: cell.fill=PatternFill("solid",fgColor=bg)
    def ajustar(ws):
        for col in ws.columns:
            mx=0; l=get_column_letter(col[0].column)
            for c in col:
                try: mx=max(mx,len(str(c.value or "")))
                except: pass
            ws.column_dimensions[l].width=min(mx+4,60)

    registros=[]
    for e in clientes:
        cliente=e.get("Cliente")
        subs=sorted(e.get("ScianSubsector",[]),key=lambda x:x.get("Percentage",0),reverse=True)
        ppal=None; secs=[]
        for i,s in enumerate(subs):
            cod=str(s.get("CodigoSubsectorSCIAN") or "N/A").strip()
            sc=cod[:2] if len(cod)>=2 else "N/A"
            info={"cod":cod,"nom":cat_sub.get(cod) or str(s.get("NameSubsectorSCIAN") or ""),
                  "sec_cod":sc,"sec_nom":cat_sec.get(sc,"Sin clasificar"),"pct":s.get("Percentage",0)}
            if i==0: ppal=info
            else: secs.append(info)
        registros.append({"cliente":cliente,"principal":ppal,"secundarias":secs})

    wb=openpyxl.Workbook(); wb.remove(wb.active)

    ws=wb.create_sheet("1. Detalle por Cliente"); ws.freeze_panes="A3"
    cols=["Cliente","Tipo Actividad","Cód. Sector","Nombre Sector","Cód. Subsector","Nombre Subsector","% Actividad"]
    for i,t in enumerate(cols,1): enc(ws.cell(row=2,column=i,value=t))
    fila=3
    for reg in registros:
        p=reg["principal"]
        if p:
            for j,v in enumerate([reg["cliente"],"PRINCIPAL",p["sec_cod"],p["sec_nom"],p["cod"],p["nom"],p["pct"]],1):
                cel(ws.cell(row=fila,column=j,value=v),bold=True,center=(j in[1,2,3,5,7]),bg=CLAR)
            fila+=1
        for s in reg["secundarias"]:
            for j,v in enumerate([reg["cliente"],"secundaria",s["sec_cod"],s["sec_nom"],s["cod"],s["nom"],s["pct"]],1):
                cel(ws.cell(row=fila,column=j,value=v),center=(j in[1,2,3,5,7]))
            fila+=1
    ws.merge_cells(start_row=fila,start_column=1,end_row=fila,end_column=7)
    c=ws.cell(row=fila,column=1,value=f"TOTAL CLIENTES: {len(registros)}")
    c.font=Font(bold=True,size=10); c.fill=PatternFill("solid",fgColor=CLAR)
    c.alignment=Alignment(horizontal="center"); c.border=borde()
    ajustar(ws)

    ws2=wb.create_sheet("2. Subsectores Principales")
    cols2=["Cód. Sector","Nombre Sector","Cód. Subsector","Nombre Subsector","Total Clientes","% del Total"]
    for i,t in enumerate(cols2,1): enc(ws2.cell(row=1,column=i,value=t))
    sub_agg=defaultdict(lambda:{"nom":"","sc":"","sc_nom":"","clientes":0})
    total=len(registros)
    for reg in registros:
        p=reg["principal"]
        if not p: continue
        sub_agg[p["cod"]]["nom"]=p["nom"]; sub_agg[p["cod"]]["sc"]=p["sec_cod"]
        sub_agg[p["cod"]]["sc_nom"]=p["sec_nom"]; sub_agg[p["cod"]]["clientes"]+=1
    for f2,(cod,d) in enumerate(sorted(sub_agg.items(),key=lambda x:-x[1]["clientes"]),2):
        pct=round(d["clientes"]/total*100,2) if total else 0
        for j,v in enumerate([d["sc"],d["sc_nom"],cod,d["nom"],d["clientes"],pct],1):
            cel(ws2.cell(row=f2,column=j,value=v),center=(j not in[2,4]))
    ajustar(ws2)

    ws3=wb.create_sheet("3. Diversificacion")
    cols3=["Cliente","Total Actividades","N° Secundarias","Subsector Principal","Sectores únicos"]
    for i,t in enumerate(cols3,1): enc(ws3.cell(row=1,column=i,value=t))
    for f3,reg in enumerate(sorted(registros,key=lambda r:len(r["secundarias"]),reverse=True),2):
        tot_act=1+len(reg["secundarias"])
        ppal_nom=(reg["principal"]["cod"]+" — "+reg["principal"]["nom"] if reg["principal"] else "—")
        secs_unic=len(set([reg["principal"]["sec_cod"]]+[s["sec_cod"] for s in reg["secundarias"]])) if reg["principal"] else 0
        for j,v in enumerate([reg["cliente"],tot_act,len(reg["secundarias"]),ppal_nom,secs_unic],1):
            cel(ws3.cell(row=f3,column=j,value=v),center=(j in[1,2,3,5]))
    ajustar(ws3)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# INIT
# ══════════════════════════════════════════════════════════════════════════════

cat_sec, cat_sub = cargar_catalogo()
clientes_raw     = cargar_json()

if not clientes_raw:
    st.error("No se encontró `datos_companyPEIBO_subsector.json` o el archivo está vacío.")
    st.stop()

df   = build_df(json.dumps(clientes_raw), cat_sec, cat_sub)
df_p = df[df["tipo_act"] == "Principal"].copy()
df_s = df[df["tipo_act"] == "Secundaria"].copy()
TOT  = df_p["cliente"].nunique()

div_df = (df.groupby("cliente")
          .agg(total_act=("sub_cod","count"),
               act_sec=("tipo_act", lambda x: (x=="Secundaria").sum()))
          .reset_index())

solo_ppal     = int((div_df["act_sec"]==0).sum())
con_sec       = int((div_df["act_sec"]>0).sum())
prom_act      = round(div_df["total_act"].mean(), 1)
max_act       = int(div_df["total_act"].max())
n_sectores    = df_p["sec_cod"].nunique()
n_subsectores = df_p["sub_cod"].nunique()

# HHI del portafolio: suma de cuadrados de la participación de cada subsector principal
hhi_base = (df_p.groupby("sub_cod").size() / TOT * 100)
HHI = int(round((hhi_base ** 2).sum()))
if HHI < 1500:   hhi_tag, hhi_color = "diversificado", C["green"]
elif HHI < 2500: hhi_tag, hhi_color = "moderado",      C["warn"]
else:            hhi_tag, hhi_color = "concentrado",   C["danger"]

db_global = (df_p.groupby(["sub_cod","sub_nom"]).size()
             .reset_index(name="Clientes").sort_values("Clientes", ascending=False))
db_global["pct"]  = (db_global["Clientes"]/TOT*100).round(1)
db_global["acum"] = db_global["pct"].cumsum().round(1)
n80 = int((db_global["acum"] <= 80).sum()) + 1

excel_bytes = generar_excel(
    json.dumps(clientes_raw), tuple(cat_sec.items()), tuple(cat_sub.items()))


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown(f"""
    <div style="padding:4px 0 18px 0;border-bottom:1px solid {C['border']};margin-bottom:20px;">
      <p style="font-size:1.1rem;font-weight:700;color:{C['text_hi']};margin:0 0 2px 0;">SCIAN Analytics</p>
      <p style="font-size:.75rem;color:{C['text_lo']};margin:0;text-transform:uppercase;
                letter-spacing:.08em;">Peibo · CrediTrust</p>
    </div>""", unsafe_allow_html=True)

    st.markdown(f'<p style="font-size:.65rem;text-transform:uppercase;letter-spacing:.1em;'
                f'color:{C["text_lo"]};margin:0 0 8px 0;">Controles</p>',
                unsafe_allow_html=True)

    top_n_ppal = st.slider(
        "Top N · Subsectores principales", 5, min(50, len(db_global)), 15,
        help="Controla el Pareto del resumen y la vista de actividad principal")

    top_n_sec = 15
    if not df_s.empty:
        db2_len = df_s.groupby("sub_cod")["cliente"].nunique().shape[0]
        top_n_sec = st.slider(
            "Top N · Subsectores secundarios", 5, min(30, db2_len), min(12, db2_len))

    st.markdown(f'<p style="font-size:.65rem;text-transform:uppercase;letter-spacing:.1em;'
                f'color:{C["text_lo"]};margin:20px 0 8px 0;">Exportar</p>',
                unsafe_allow_html=True)
    st.download_button(
        label="Descargar reporte Excel",
        data=excel_bytes,
        file_name="reporte_scian_peibo.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown(f"""
    <div style="margin-top:28px;padding-top:14px;border-top:1px solid {C['border']};">
      <p style="font-size:.67rem;color:{C['text_lo']};line-height:1.6;margin:0;">
        Clasificación SCIAN 2023.<br>
        Reemplazar <code style="color:{C['text_md']};">datos_companyPEIBO_subsector.json</code><br>
        para actualizar datos.
      </p>
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE HEADER
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(f"""
<div class="ph-wrap">
  <p class="ph-eyebrow">Análisis de portafolio · SCIAN 2023</p>
  <h1 class="ph-title">Actividades Económicas — Peibo</h1>
  <p class="ph-sub">Clasificación industrial del portafolio de clientes según el
  Sistema de Clasificación Industrial de América del Norte.</p>
</div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════

tab_resumen, tab_act, tab_div = st.tabs([
    "Resumen Global",
    "Actividad Económica",
    "Diversificación",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — RESUMEN GLOBAL
# ══════════════════════════════════════════════════════════════════════════════

with tab_resumen:

    cards = (
        kpi_card(f"{TOT:,}",        "Total clientes",       "Portafolio completo",                           "kc-gold")
      + kpi_card(f"{n_sectores}",   "Sectores únicos",      "Clasificación macro SCIAN",                     "kc-blue")
      + kpi_card(f"{n_subsectores}","Subsectores únicos",   "Clasificación micro SCIAN",                     "kc-indigo")
      + kpi_card(f"{con_sec:,}",    "Diversificados",       f"{round(con_sec/TOT*100,1)}% del portafolio",   "kc-green")
      + kpi_card(f"{prom_act}",     "Actividades / cliente","Máximo registrado: " + str(max_act),            "kc-warn")
      + kpi_card(f"{HHI:,}",        "Índice HHI",           hhi_tag.capitalize(),                            "kc-purple")
    )
    kpi_grid(cards)

    sec_dom = df_p.groupby("sec_cod").size().idxmax()
    st.markdown(f"""
    <div class="insight-box">
      <b>{n80} subsectores</b> concentran el 80% del portafolio.
      El sector de mayor frecuencia es <b>{sec_dom}</b>.
      <b>{round(con_sec/TOT*100,1)}%</b> de los clientes opera en más de un sector.
      Índice de concentración HHI: <b>{HHI:,}</b> ({hhi_tag}).
    </div>""", unsafe_allow_html=True)

    sec("Jerarquía Sectorial y Curva de Concentración",
        "Vista panorámica del portafolio: composición por sector/subsector y "
        "qué tan concentrada está la actividad económica en pocos subsectores.")

    col_sun, col_par = st.columns([1, 1.15])

    with col_sun:
        with chart_card():
            chart_header("Jerarquía sectorial",
                         "Anillo interno = sector. Anillo externo = subsector dentro de ese sector. "
                         "Pasa el mouse sobre un segmento para ver el detalle — el texto se omite "
                         "a propósito para mantener la vista legible con muchos sectores.")

            top_n_sun = st.slider(
                "Top N · Sectores en el sunburst", 3, min(15, n_sectores), min(8, n_sectores),
                key="sl_sunburst",
                help="Limita la jerarquía a los N sectores con más clientes; el resto se agrupa en \"Otros\"")

            # Conteo único de clientes por sector/subsector — un cliente cuenta una sola vez
            sun_base = (df_p.groupby(["sec_cod", "sec_nom", "sub_cod", "sub_nom"])["cliente"]
                        .nunique().reset_index(name="n_clientes"))

            sec_totales = (sun_base.groupby(["sec_cod", "sec_nom"])["n_clientes"]
                           .sum().sort_values(ascending=False))
            top_secs = sec_totales.head(top_n_sun).index.get_level_values("sec_cod").tolist()

            sun_top = sun_base[sun_base["sec_cod"].isin(top_secs)].copy()
            sun_rest = sun_base[~sun_base["sec_cod"].isin(top_secs)].copy()

            if not sun_rest.empty:
                otros_total = sun_rest["n_clientes"].sum()
                otros_row = pd.DataFrame([{
                    "sec_cod": "OTR", "sec_nom": "Otros sectores",
                    "sub_cod": "OTR", "sub_nom": f"{sun_rest['sec_cod'].nunique()} sectores agrupados",
                    "n_clientes": otros_total,
                }])
                sun_top = pd.concat([sun_top, otros_row], ignore_index=True)

            sun_top["sec_label"] = sun_top["sec_cod"] + " · " + sun_top["sec_nom"]
            sun_top["sub_label"] = sun_top["sub_cod"] + " · " + sun_top["sub_nom"]

            palette = [C["blue"], C["indigo"], C["green"], C["gold"],
                       C["purple"], C["warn"], C["blue_mid"], C["indigo2"]]
            n_unique_secs = sun_top["sec_label"].nunique()
            color_map = {lbl: palette[i % len(palette)]
                         for i, lbl in enumerate(sun_top["sec_label"].unique())}
            if "OTR · Otros sectores" in color_map:
                color_map["OTR · Otros sectores"] = C["text_dim"]

            fig_sun = px.sunburst(
                sun_top, path=["sec_label", "sub_label"], values="n_clientes",
                color="sec_label", color_discrete_map=color_map,
            )
            fig_sun.update_traces(
                textinfo="none",
                marker=dict(line=dict(color=C["bg"], width=2)),
                hovertemplate="<b>%{label}</b><br>Clientes: <b>%{value}</b><br>%{percentParent:.1%} de su grupo<extra></extra>",
            )
            lay(fig_sun, margin=dict(l=4,r=4,t=4,b=4), height=380)
            pch(fig_sun)

            if not sun_rest.empty:
                st.caption(
                    f"\"Otros sectores\" agrupa {sun_rest['sec_cod'].nunique()} sectores adicionales "
                    f"con {int(sun_rest['n_clientes'].sum()):,} clientes en total — "
                    f"aumenta el Top N para verlos desagregados."
                )

    with col_par:
        with chart_card():
            chart_header("Curva de Pareto global",
                         "Barras = clientes por subsector. Línea ámbar = % acumulado del portafolio.",
                         f"Top {top_n_ppal} subsectores — ajustable en el panel izquierdo")

            dt_p   = db_global.head(top_n_ppal)
            cats_p = dt_p["sub_cod"].astype(str).tolist()

            fig_p = go.Figure()
            fig_p.add_trace(go.Bar(
                x=cats_p, y=dt_p["Clientes"], name="Clientes",
                marker=dict(color=dt_p["Clientes"],
                            colorscale=[[0, C["blue_dim"]], [0.45, C["blue_mid"]], [1, C["blue"]]],
                            line=dict(width=0)),
                customdata=list(zip(dt_p["sub_nom"], dt_p["pct"])),
                hovertemplate="<b>%{x}</b> — %{customdata[0]}<br>Clientes: <b>%{y}</b> · %{customdata[1]}%<extra></extra>",
            ))
            fig_p.add_trace(go.Scatter(
                x=cats_p, y=dt_p["acum"], name="% Acumulado", yaxis="y2",
                line=dict(color=C["warn"], width=2.5), mode="lines+markers",
                marker=dict(size=5, color=C["warn"]),
                hovertemplate="%{y:.1f}% acumulado<extra></extra>",
            ))
            fig_p.add_shape(type="line",
                x0=-0.5, x1=len(cats_p)-0.5, y0=80, y1=80, yref="y2",
                line=dict(color=C["danger"], dash="dot", width=1.5))
            fig_p.add_annotation(
                x=len(cats_p)-1, y=80, yref="y2", text="80%",
                showarrow=False, font=dict(color=C["danger"], size=10),
                xanchor="right", yanchor="bottom")
            lay(fig_p,
                xaxis=dict(type="category", tickangle=-40, **GR, tickfont=dict(size=9)),
                yaxis=dict(title="Clientes", **GR),
                yaxis2=dict(overlaying="y", side="right", range=[0,108],
                            ticksuffix="%", gridcolor="rgba(0,0,0,0)",
                            tickfont=dict(color=C["text_lo"], size=10)),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, x=0, font=dict(size=10)),
                barmode="overlay",
                height=380)
            pch(fig_p)

    hr()
    sec("Registro Detallado de Actividades por Cliente",
        "Cada fila es una actividad registrada. La barra de progreso refleja el peso "
        "relativo de esa actividad dentro del perfil del cliente.")

    st.dataframe(
        df[["cliente","tipo_act","sec_cod","sec_nom","sub_cod","sub_nom","pct"]]
        .rename(columns={
            "cliente":  "ID Cliente",
            "tipo_act": "Jerarquía",
            "sec_cod":  "Cód. Sector",
            "sec_nom":  "Sector",
            "sub_cod":  "Cód. Subsector",
            "sub_nom":  "Subsector",
            "pct":      "% Operación",
        })
        .sort_values(["ID Cliente","% Operación"], ascending=[True,False]),
        use_container_width=True, hide_index=True, height=360,
        column_config={
            "% Operación": st.column_config.ProgressColumn(
                format="%d%%", min_value=0, max_value=100),
        }
    )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — ACTIVIDAD ECONÓMICA (principal + secundaria unificadas)
# ══════════════════════════════════════════════════════════════════════════════

with tab_act:

    sec("Sectores: Actividad Principal vs Secundaria",
        "Comparación directa por sector — sin cambiar de pestaña.")

    # ── Datos para la barra divergente ──────────────────────────────────────
    ds_p = (df_p.groupby(["sec_cod","sec_nom"]).size()
            .reset_index(name="Principal"))
    ds_s = (df_s.groupby(["sec_cod","sec_nom"])["cliente"].nunique()
            .reset_index(name="Secundaria")) if not df_s.empty else \
           pd.DataFrame(columns=["sec_cod","sec_nom","Secundaria"])

    ds_div = ds_p.merge(ds_s[["sec_cod","Secundaria"]], on="sec_cod", how="left")
    ds_div["Secundaria"] = ds_div["Secundaria"].fillna(0).astype(int)
    ds_div["sec_label"]  = ds_div["sec_cod"] + " " + ds_div["sec_nom"]

    # índice de asimetría: ratio entre el lado dominante y el lado menor (mín 1 para evitar /0)
    def asim_ratio(row):
        p, s = row["Principal"], row["Secundaria"]
        if p == 0 and s == 0: return 0.0
        if s == 0: return float(p)  # solo principal → asimetría máxima hacia ese lado
        if p == 0: return float(s)
        return max(p, s) / max(min(p, s), 1)

    ds_div["asimetria"] = ds_div.apply(asim_ratio, axis=1)
    ds_div["lado_fuerte"] = ds_div.apply(
        lambda r: "principal" if r["Principal"] >= r["Secundaria"] else "secundaria", axis=1)
    ds_div = ds_div.sort_values("asimetria", ascending=False)

    max_lado = max(ds_div["Principal"].max(), ds_div["Secundaria"].max(), 1)

    # Hallazgo automático — sector más asimétrico hacia secundaria (si existe)
    hacia_sec = ds_div[(ds_div["lado_fuerte"] == "secundaria") & (ds_div["Secundaria"] > 0)]
    if not hacia_sec.empty:
        top_asim = hacia_sec.iloc[0]
        insight_txt = (
            f"<b>{top_asim['sec_nom']}</b> es <b>{top_asim['asimetria']:.1f}x</b> más frecuente "
            f"como actividad secundaria que como principal — posible señal de diversificación "
            f"hacia ese sector."
        )
    else:
        top_asim_row = ds_div.iloc[0]
        insight_txt = (
            f"Todos los sectores son más fuertes como actividad principal. "
            f"El más asimétrico es <b>{top_asim_row['sec_nom']}</b> "
            f"({top_asim_row['asimetria']:.1f}x)."
        )

    with chart_card():
        chart_header(
            "Sectores: principal vs secundaria",
            "Cada fila es un sector. La barra <em>azul</em> (derecha) muestra cuántos clientes "
            "tienen ese sector como su actividad principal. La barra <i>índigo</i> (izquierda) "
            "muestra cuántos lo tienen como secundaria.",
            "Ordenado por índice de asimetría — pasa el mouse sobre un segmento para ver el detalle exacto",
        )

        legend_row([("Actividad principal", C["blue"]), ("Actividad secundaria", C["indigo"])])

        st.markdown(f"""
        <div class="insight-box">{insight_txt}</div>
        """, unsafe_allow_html=True)

        cats_div = ds_div["sec_label"].tolist()

        fig_div = go.Figure()
        fig_div.add_trace(go.Bar(
            x=-ds_div["Secundaria"], y=cats_div, orientation="h", name="Secundaria",
            marker=dict(color=C["indigo"], line=dict(width=0)),
            customdata=list(zip(ds_div["sec_nom"], ds_div["Secundaria"], ds_div["Principal"])),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Como secundaria: <b>%{customdata[1]}</b> clientes<br>"
                "Como principal: %{customdata[2]} clientes"
                "<extra></extra>"
            ),
            text=ds_div["Secundaria"].astype(str),
            textposition="outside",
            textfont=dict(color=C["text_lo"], size=10),
        ))
        fig_div.add_trace(go.Bar(
            x=ds_div["Principal"], y=cats_div, orientation="h", name="Principal",
            marker=dict(color=C["blue"], line=dict(width=0)),
            customdata=list(zip(ds_div["sec_nom"], ds_div["Principal"], ds_div["Secundaria"])),
            hovertemplate=(
                "<b>%{customdata[0]}</b><br>"
                "Como principal: <b>%{customdata[1]}</b> clientes<br>"
                "Como secundaria: %{customdata[2]} clientes"
                "<extra></extra>"
            ),
            text=ds_div["Principal"].astype(str),
            textposition="outside",
            textfont=dict(color=C["text_lo"], size=10),
        ))
        fig_div.add_vline(x=0, line=dict(color=C["border2"], width=1, dash="dot"))
        lay(fig_div,
            barmode="overlay",
            xaxis=dict(
                range=[-max_lado*1.25, max_lado*1.25],
                tickvals=[-max_lado, 0, max_lado],
                ticktext=["Secundaria", "Balance", "Principal"],
                **GR,
            ),
            yaxis=dict(type="category", categoryorder="array", categoryarray=cats_div[::-1],
                       **GR, tickfont=dict(size=10, color=C["text_md"])),
            showlegend=False,
            height=max(320, len(ds_div)*42+90))
        pch(fig_div)

    hr()
    sec("Subsectores por Tipo de Actividad",
        "Desagregación a nivel subsector, separada por tipo. "
        "Los sliders del panel izquierdo controlan cada lista de forma independiente.")

    col_sp, col_ss = st.columns([1, 1])

    with col_sp:
        with chart_card():
            chart_header(f"Top {top_n_ppal} subsectores — principal",
                         "Subsectores con más clientes en actividad principal")
            dt_ppal = db_global.head(top_n_ppal)
            pch(bar_h(dt_ppal, "Clientes", "sub_cod", "sub_nom", "pct",
                      C["blue_dim"], C["blue"]))

    with col_ss:
        with chart_card():
            if df_s.empty:
                chart_header("Top subsectores — secundaria",
                             "No hay actividades secundarias registradas en este portafolio")
                st.info("Sin datos de actividad secundaria.")
            else:
                db2 = (df_s.groupby(["sub_cod","sub_nom"])["cliente"].nunique()
                       .reset_index(name="Clientes").sort_values("Clientes", ascending=False))
                db2["pct"] = (db2["Clientes"]/TOT*100).round(1)
                chart_header(f"Top {top_n_sec} subsectores — secundaria",
                             "Subsectores con más clientes en actividad complementaria")
                pch(bar_h(db2.head(top_n_sec), "Clientes", "sub_cod", "sub_nom", "pct",
                          C["indigo_dim"], C["indigo"]))


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — DIVERSIFICACIÓN
# ══════════════════════════════════════════════════════════════════════════════

with tab_div:

    div_cards = (
        kpi_card(f"{solo_ppal:,}", "Solo act. principal",   f"{round(solo_ppal/TOT*100,1)}% del portafolio", "kc-slate")
      + kpi_card(f"{con_sec:,}",   "Con act. secundarias",  f"{round(con_sec/TOT*100,1)}% del portafolio",   "kc-green")
      + kpi_card(f"{prom_act}",    "Actividades / cliente", "Promedio del portafolio",                       "kc-warn")
      + kpi_card(f"{max_act}",     "Máx. actividades",      "Cliente más diversificado",                     "kc-danger")
    )
    kpi_grid(div_cards, cols=4)

    sec("Distribución de Actividades por Cliente",
        "Cuántos clientes tienen 1, 2, 3… actividades registradas. "
        "Un cliente con 1 actividad opera en un único sector.")

    col_g, col_h = st.columns([1, 1])

    with col_g:
        with chart_card():
            dist = (div_df["total_act"].value_counts()
                    .reset_index()
                    .rename(columns={"total_act":"N actividades","count":"Clientes"})
                    .sort_values("N actividades"))
            dist["pct"] = (dist["Clientes"]/TOT*100).round(1)

            chart_header("Distribución de frecuencia",
                         "Cada barra representa un número de actividades distinto")
            fig = go.Figure(go.Bar(
                x=dist["N actividades"].astype(str),
                y=dist["Clientes"],
                marker=dict(color=dist["Clientes"],
                            colorscale=[[0, C["blue_dim"]], [1, C["blue"]]],
                            line=dict(width=0)),
                customdata=dist["pct"],
                text=dist["Clientes"],
                textposition="outside",
                textfont=dict(color=C["text_lo"], size=11),
                hovertemplate=(
                    "<b>%{x} actividades</b><br>"
                    "Clientes: <b>%{y}</b> · %{customdata}% del portafolio"
                    "<extra></extra>"
                ),
            ))
            for _, row in dist.iterrows():
                fig.add_annotation(
                    x=str(row["N actividades"]), y=row["Clientes"],
                    text=f"{row['pct']}%", showarrow=False,
                    yshift=18, font=dict(size=9, color=C["text_md"]))
            lay(fig,
                xaxis=dict(title="N° de actividades registradas", type="category", **GR),
                yaxis=dict(title="N° de clientes", **GR),
                height=320)
            pch(fig)

    with col_h:
        with chart_card():
            top_n_div_quick = st.slider(
                "Top N · Clientes más diversificados", 5, min(30, len(div_df)), min(12, len(div_df)),
                key="sl_topdiv", help="Cantidad de clientes a mostrar en el ranking")

            top_div = (div_df.sort_values("total_act", ascending=False)
                       .head(top_n_div_quick)
                       .rename(columns={"cliente": "Cliente", "total_act": "Total", "act_sec": "Secundarias"}))
            top_div["Cliente_label"] = top_div["Cliente"].astype(str).apply(
                lambda v: v if v.strip().lower().startswith("cliente") else f"Cliente {v}")

            chart_header(f"Top {top_n_div_quick} clientes diversificados",
                         "Ordenados por número total de actividades registradas")

            fig_topdiv = go.Figure(go.Bar(
                x=top_div["Total"], y=top_div["Cliente_label"], orientation="h",
                marker=dict(color=top_div["Total"],
                            colorscale=[[0, C["blue_dim"]], [1, C["blue"]]],
                            line=dict(width=0)),
                customdata=top_div["Secundarias"],
                text=top_div["Total"], textposition="outside",
                textfont=dict(color=C["text_lo"], size=10),
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "Total actividades: <b>%{x}</b><br>"
                    "Secundarias: <b>%{customdata}</b>"
                    "<extra></extra>"
                ),
            ))
            lay(fig_topdiv,
                xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                yaxis=dict(type="category", categoryorder="total ascending",
                           showgrid=False, tickfont=dict(color=C["text_md"], size=10)),
                height=max(280, top_n_div_quick*28+80))
            pch(fig_topdiv)

    hr()
    sec("Mapa de Concentración y Combinaciones",
        "El mapa de calor muestra el patrón completo de actividades por cliente, "
        "limitado a los subsectores que realmente participan en combinaciones "
        "principal → secundaria. Así ambas piezas siempre describen el mismo universo.")

    # ── Combos reales primero (sin restricción) — el heatmap se deriva de ellos ──
    combos = get_combos(df_p, df_s, top=10)

    if combos.empty:
        st.info(
            "Este portafolio no registra clientes con actividad secundaria que "
            "forme combinaciones con su actividad principal — todos los clientes "
            "diversificados tienen secundarias que no se repiten como patrón."
        )
    else:
        # Universo de subsectores = los que aparecen en los combos reales
        universo_subs = sorted(set(combos["ppal_cod"]) | set(combos["sec_cod"]))

        df_universe = df[df["sub_cod"].isin(universo_subs)]
        clientes_relevantes = (df_universe.groupby("cliente")["sub_cod"].nunique()
                                .sort_values(ascending=False).head(8).index.tolist())

        with chart_card():
            chart_header(
                "Mapa de calor cliente × subsector",
                f"Filas = clientes con más actividad dentro de los {len(universo_subs)} "
                "subsectores que participan en combinaciones (abajo). "
                "Columnas = esos mismos subsectores.",
                "Color más claro = mayor % de participación de esa actividad para el cliente",
            )

            heat_df = df_universe[df_universe["cliente"].isin(clientes_relevantes)]
            pivot = heat_df.pivot_table(
                index="cliente", columns="sub_cod", values="pct",
                aggfunc="max", fill_value=0
            ).reindex(columns=universo_subs, fill_value=0)
            pivot = pivot.reindex(clientes_relevantes)

            cliente_labels = [
                v if str(v).strip().lower().startswith("cliente") else f"Cliente {v}"
                for v in pivot.index
            ]
            sub_nom_map = dict(zip(df_universe["sub_cod"], df_universe["sub_nom"]))
            col_labels = [f"{c} · {sub_nom_map.get(c,'')[:18]}" for c in pivot.columns]

            fig_heat = go.Figure(go.Heatmap(
                z=pivot.values,
                x=col_labels,
                y=cliente_labels,
                colorscale=[[0, C["surface2"]], [0.4, C["blue_dim"]], [0.75, C["blue_mid"]], [1, C["blue"]]],
                zmin=0, zmax=100,
                hovertemplate="<b>%{y}</b><br>%{x}<br>Participación: <b>%{z:.0f}%</b><extra></extra>",
                colorbar=dict(
                    title=dict(text="% Actividad", font=dict(color=C["text_md"], size=10)),
                    tickfont=dict(color=C["text_md"], size=9),
                    len=0.7, thickness=10,
                ),
                xgap=3, ygap=3,
            ))
            lay(fig_heat,
                xaxis=dict(tickfont=dict(size=9, color=C["text_md"]), tickangle=-20, side="bottom"),
                yaxis=dict(tickfont=dict(size=10, color=C["text_md"]), autorange="reversed"),
                height=max(260, len(cliente_labels)*36+80))
            pch(fig_heat)

        st.markdown(
            f"""<div style="display:flex;justify-content:center;margin:2px 0 14px 0;">
                <svg width="20" height="22" viewBox="0 0 20 22">
                  <path d="M10 2 L10 16 M4 12 L10 18 L16 12" fill="none"
                        stroke="{C['green']}" stroke-width="1.5"
                        stroke-linecap="round" stroke-linejoin="round"/>
                </svg></div>""",
            unsafe_allow_html=True)

        with chart_card():
            chart_header(
                "Combinaciones principal → secundaria",
                "Pares de subsectores que co-ocurren con más frecuencia en el portafolio.",
                "Estos pares definen el universo de subsectores mostrado en el mapa de calor, arriba",
            )
            combos["combo"]   = combos["ppal_cod"] + "  →  " + combos["sec_cod"]
            combos["detalle"] = combos["ppal_nom"] + "  →  " + combos["sec_nom"]
            fig_combo = go.Figure(go.Bar(
                x=combos["Clientes"], y=combos["combo"], orientation="h",
                marker=dict(color=combos["Clientes"],
                            colorscale=[[0, C["indigo_dim"]], [1, C["indigo"]]],
                            line=dict(width=0)),
                customdata=combos["detalle"],
                text=combos["Clientes"], textposition="outside",
                textfont=dict(color=C["text_lo"], size=10),
                hovertemplate=(
                    "<b>%{customdata}</b><br>"
                    "Clientes con esta combinación: <b>%{x}</b>"
                    "<extra></extra>"
                ),
            ))
            lay(fig_combo,
                xaxis=dict(showgrid=False, showticklabels=False),
                yaxis=dict(type="category", categoryorder="total ascending",
                           showgrid=False, tickfont=dict(size=10, color=C["text_md"])),
                height=max(280, len(combos)*36+60))
            pch(fig_combo)
